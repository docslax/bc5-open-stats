import json
from pathlib import Path

import openpyxl

WORKBOOK = Path("external/2026BCStats final (2).xlsx")


def decode_bowler_id(bowler_id):
    """Decode the structured bowler identifier used in the workbook."""
    try:
        value = int(bowler_id)
    except (TypeError, ValueError):
        return None

    zone = value // 100
    category_code = (value % 100) // 10
    qualifier = value % 10

    category = {2: "Mens", 3: "Ladies", 4: "Mixed"}.get(category_code, "Unknown")
    return {
        "bowlerId": value,
        "zone": zone,
        "categoryCode": category_code,
        "category": category,
        "qualifier": qualifier,
        "isCoach": qualifier == 0,
    }


def summarize_sheet(ws, limit=8):
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=min(limit, ws.max_row), values_only=False):
        vals = []
        for cell in r:
            value = cell.value
            if value is None:
                vals.append(None)
            else:
                vals.append(str(value) if isinstance(value, (int, float, str)) else repr(value))
        rows.append(vals)
    return {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "preview": rows,
    }


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    summary = {
        "workbook": str(WORKBOOK),
        "sheet_names": wb.sheetnames,
        "sheets": [],
        "candidate_standings_sheets": [],
    }

    for ws in wb.worksheets:
        entry = summarize_sheet(ws, limit=8)
        summary["sheets"].append(entry)

        if ws.title == "CODES":
            sample_ids = []
            for row in ws.iter_rows(min_row=2, max_row=min(12, ws.max_row), values_only=True):
                bowler_id = row[0]
                if isinstance(bowler_id, (int, float)):
                    sample_ids.append({"raw": int(bowler_id), **decode_bowler_id(int(bowler_id))})
            summary["sample_bowler_ids"] = sample_ids

        title = ws.title.lower()
        if any(token in title for token in ("ladder", "points", "indstats", "step")):
            summary["candidate_standings_sheets"].append({
                "name": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
            })

    out = Path("external/workbook_summary.json")
    out.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
